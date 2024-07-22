import cv2
import mediapipe as mp
import numpy as np
import pandas as pd
import os
import easygui
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image 
from openpyxl.utils.dataframe import dataframe_to_rows 
from openpyxl.chart import Reference
from openpyxl.chart import BarChart


mp_pose = mp.solutions.pose
pose_estimator = mp_pose.Pose(static_image_mode=False, min_detection_confidence=0.5, min_tracking_confidence=0.5)

df1 = pd.DataFrame(columns=['Frame', 'Landmark', 'Name', 'X', 'Y', 'Z', 'Visibility'])
df2 = pd.DataFrame(columns=['Frame', 'Landmark', 'Name', 'X', 'Y', 'Z', 'Visibility'])

file_paths = easygui.fileopenbox(title='Select Cricket Shot Images/Videos', multiple=True)


excel_file_path = f'cricket_shot_data_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
workbook = Workbook()


for file_path in file_paths:
   
    is_image = file_path.lower().endswith(('.jpg', '.jpeg', '.png'))
    is_video = file_path.lower().endswith('.mp4')
    
    if is_image:
        frames = [cv2.imread(file_path)]
        frame_counts = ['Image']
    elif is_video:
        cap = cv2.VideoCapture(file_path)
        frames = []
        frame_counts = []
        frame_count = 0
        ret, frame = cap.read()
        while ret:
            frames.append(frame)
            frame_counts.append(frame_count)
            ret, frame = cap.read()
            frame_count += 1
        cap.release()
    else:
        continue
    
    sheet_name = os.path.splitext(os.path.basename(file_path))[0]
    worksheet = workbook.create_sheet(title=sheet_name)
    
    for frame_count, frame in enumerate(frames):
       
        annotated_image_path = f'annotated_{frame_count}.png'
        cv2.imwrite(annotated_image_path, frame)
        
      
        img = Image(annotated_image_path)
        worksheet.add_image(img, f'A{frame_count + 2}')
        
        os.remove(annotated_image_path)
        
       
        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        
        
        results = pose_estimator.process(frame_rgb)
        
     
        if results.pose_landmarks is not None:
           
            landmarks = results.pose_landmarks.landmark
           
            for i, landmark in enumerate(landmarks):
                x = landmark.x
                y = landmark.y
                z = landmark.z
                visibility = landmark.visibility
              
                if file_path == file_paths[0]:
                    df1 = df1._append({'Frame': frame_count, 'Landmark': i, 'Name': mp_pose.PoseLandmark(i).name, 'X': x, 'Y': y, 'Z': z, 'Visibility': visibility},
                                     ignore_index=True)
                elif file_path == file_paths[1]:
                    df2 = df2._append({'Frame': frame_count, 'Landmark': i, 'Name': mp_pose.PoseLandmark(i).name, 'X': x, 'Y': y, 'Z': z, 'Visibility': visibility},
                                     ignore_index=True)

worksheet1 = workbook.active
worksheet1.title = 'Data1'
for row in dataframe_to_rows(df1, index=True, header=True):
    worksheet1.append(row)

worksheet2 = workbook.create_sheet(title='Data2')
for row in dataframe_to_rows(df2, index=True, header=True):
    worksheet2.append(row)

comparison_sheet = workbook.create_sheet(title='Comparison')

data1_x = Reference(worksheet=worksheet1, min_col=6, min_row=2, max_row=len(df1) + 1)
data1_y = Reference(worksheet=worksheet1, min_col=7, min_row=2, max_row=len(df1) + 1)
data1_z = Reference(worksheet=worksheet1, min_col=8, min_row=2, max_row=len(df1) + 1)
data1_visibility = Reference(worksheet=worksheet1, min_col=9, min_row=2, max_row=len(df1) + 1)
data2_x = Reference(worksheet=worksheet2, min_col=6, min_row=2, max_row=len(df2) + 1)
data2_y = Reference(worksheet=worksheet2, min_col=7, min_row=2, max_row=len(df2) + 1)
data2_z = Reference(worksheet=worksheet2, min_col=8, min_row=2, max_row=len(df2) + 1)
data2_visibility = Reference(worksheet=worksheet2, min_col=9, min_row=2, max_row=len(df2) + 1)
categories = Reference(worksheet=worksheet1, min_col=5, min_row=2, max_row=len(df1) + 1)

chart_x = BarChart()
chart_x.title = 'Comparison of X Landmarks'
chart_x.x_axis.title = 'Landmark'
chart_x.y_axis.title = 'X'
chart_x.add_data(data1_x, titles_from_data=True)
chart_x.add_data(data2_x, titles_from_data=True)
chart_x.set_categories(categories)
comparison_sheet.add_chart(chart_x, 'A1')

chart_y = BarChart()
chart_y.title = 'Comparison of Y Landmarks'
chart_y.x_axis.title = 'Landmark'
chart_y.y_axis.title = 'Y'
chart_y.add_data(data1_y, titles_from_data=True)
chart_y.add_data(data2_y, titles_from_data=True)
chart_y.set_categories(categories)
comparison_sheet.add_chart(chart_y, 'A18')

chart_z = BarChart()
chart_z.title = 'Comparison of Z Landmarks'
chart_z.x_axis.title = 'Landmark'
chart_z.y_axis.title = 'Z'
chart_z.add_data(data1_z, titles_from_data=True)
chart_z.add_data(data2_z, titles_from_data=True)
chart_z.set_categories(categories)
comparison_sheet.add_chart(chart_z, 'A35')

chart_visibility = BarChart()
chart_visibility.title = 'Comparison of Visibility'
chart_visibility.x_axis.title = 'Landmark'
chart_visibility.y_axis.title = 'Visibility'
chart_visibility.add_data(data1_visibility, titles_from_data=True)
chart_visibility.add_data(data2_visibility, titles_from_data=True)
chart_visibility.set_categories(categories)
comparison_sheet.add_chart(chart_visibility, 'A52')


image1 = cv2.imread(file_paths[0])
image2 = cv2.imread(file_paths[1])
diff_image = cv2.absdiff(image1, image2)
diff_gray = cv2.cvtColor(diff_image, cv2.COLOR_BGR2GRAY)
_, diff_thresh = cv2.threshold(diff_gray, 30, 255, cv2.THRESH_BINARY)
contours, _ = cv2.findContours(diff_thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
cv2.drawContours(image1, contours, -1, (0, 0, 255), 2)
diff_image_path = f'diff_image.png'
cv2.imwrite(diff_image_path, image1)
diff_img = Image(diff_image_path)
comparison_sheet.add_image(diff_img, 'A70')
os.remove(diff_image_path)


workbook.save(excel_file_path)
print(f"Data saved to {excel_file_path} file with visualization, comparison, and highlighted differences.")